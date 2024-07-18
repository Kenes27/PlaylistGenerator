import os
import sys
import random
import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter import *
from pydub import AudioSegment
import json


def find_ab(percentage):
    if percentage == 0:
        return 0
    elif 0.05 >= percentage > 0:
        return 1
    elif 0.14 >= percentage > 0.05:
        return 2
    elif 0.19 >= percentage > 0.14:
        return 3
    elif 1 >= percentage > 0.14:
        return 4
    else:
        print("too much")
        return -1


def hour_to_seconds(time):
    a = 0
    b = ''
    index = 2
    for x in time:
        if ord(x) == ord(':'):
            a += int(b) * pow(60, index)
            index -= 1
            b = ''
            continue
        b += str(x)
    a += int(b)
    return a


def sec_to_hour(time):
    a = int(time / 3600)
    b = ''
    if len(str(a)) == 1:
        b = "0" + b + str(a) + ':'
    else:
        b = b + str(a) + ':'
    time -= a * 3600

    a = int(time / 60)
    if len(str(a)) == 1:
        b = b + '0' + str(a) + ':'
    else:
        b = b + str(a) + ':'

    time -= a * 60
    if len(str(time)) == 1:
        b = b + '0' + str(time)
    else:
        b = b + str(time)
    return b


def sort_list(repeat):
    pos = 0
    iter = 0
    arr_llst = []
    num = [5, 10, 15, 20]
    for i in range(len(repeat)):
        arr_llst.append(i)
    for j in num:
        for i in repeat[pos:len(repeat)]:
            if i == j:
                if pos == iter:
                    iter += 1
                    pos += 1
                    continue
                b = arr_llst[pos]
                arr_llst[pos] = arr_llst[iter]
                arr_llst[iter] = b
                a = repeat[pos]
                repeat[pos] = repeat[iter]
                repeat[iter] = a
                pos += 1
            iter += 1
        iter = pos
    return arr_llst, repeat


def rearrange(index, lists):
    new_list = []
    for i in range(len(index)):
        new_list.append(lists[index[i]])
    return new_list


class MediaPlanApp:

    def __init__(self, root):
        self.root = root
        self.root.title("Генератор медиаплана")

        # Инициализация переменных
        self.ad_dur = []
        self.ad_repeat = []
        self.ad_name = []
        self.ad_files = []
        self.music_files = []
        self.ad_frame_list = []
        self.number = 0

        # Настройка начального интерфейса
        self.setup_initial_gui()

    def setup_initial_gui(self):

        exe_path = os.path.abspath(sys.argv[0])
        exe_dir = os.path.dirname(exe_path)

        json_data = None

        if os.path.isfile(exe_dir + '/data.json'):
            with open(exe_dir + '/data.json', encoding='ascii') as f:
                json_data = json.load(f)

        self.frame = tk.Frame(self.root, padx=10, pady=10)
        self.frame.pack(padx=10, pady=10)

        tk.Label(self.frame, text="Папка с музыкальными файлами:", anchor='w').grid(row=0, column=0, sticky=tk.W,
                                                                                    padx=5, pady=5)
        self.music_files_entry = tk.Entry(self.frame, width=50)
        self.music_files_entry.grid(row=0, column=1, padx=5, pady=5)
        if json_data != None:
            self.music_files_entry.insert(0, json_data["music"])
        tk.Button(self.frame, text="Обзор", command=self.browse_music_files).grid(row=0, column=2, padx=5, pady=5)

        tk.Label(self.frame, text="Начальное время:", anchor='w').grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.start_time_entry = tk.Entry(self.frame, width=10)
        if json_data != None and json_data["start"] != "":
            self.start_time_entry.insert(0, json_data["start"])
        else:
            self.start_time_entry.insert(0, "09:00:00")
        self.start_time_entry.grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
        self.start_time_entry.bind("<KeyRelease>", lambda event: self.update_load())

        tk.Label(self.frame, text="Конечное время:", anchor='w').grid(row=1, column=1, sticky=tk.W, padx=5, pady=5)
        self.end_time_entry = tk.Entry(self.frame, width=10)
        if json_data != None and json_data["end"] != "":
            self.end_time_entry.insert(0, json_data["end"])
        else:
            self.end_time_entry.insert(0, "16:00:00")
        self.end_time_entry.grid(row=2, column=1, sticky=tk.W, padx=5, pady=5)
        self.end_time_entry.bind("<KeyRelease>", lambda event: self.update_load())

        self.load = tk.Label(self.frame, text="", anchor='w')
        self.load.grid(row=1, column=2, sticky=tk.W, padx=5, pady=5)

        tk.Button(self.frame, text="Добавить рекламу", command=self.add_advertisement).grid(row=3, column=0, pady=10)

        # Create a canvas for scrolling
        self.canvas = Canvas(self.root, borderwidth=0, width=500, height=300)
        self.ad_frame = tk.Frame(self.canvas, padx=10, pady=10)
        self.scrollbar = Scrollbar(self.root, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.scrollbar.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.canvas.create_window((4, 4), window=self.ad_frame, anchor="nw", tags="self.ad_frame")

        self.ad_frame.bind("<Configure>", lambda event, canvas=self.canvas: self.on_frame_configure(canvas))

        tk.Button(self.frame, text="Удалить рекламу", command=self.delete_advertisement).grid(row=3, column=1, pady=10)
        tk.Button(self.frame, text="Сгенерировать медиаплан", command=self.generate_media_plan).grid(row=3, column=2,
                                                                                                     pady=10)

        if json_data != None and "ad" in json_data.keys():
            for x in json_data["ad"]:
                self.add_advertisement_init(x[0], x[1])

    def on_frame_configure(self, canvas):
        canvas.configure(scrollregion=canvas.bbox("all"))

    def browse_music_files(self):
        files = filedialog.askdirectory()
        if files:
            self.music_files_entry.delete(0, tk.END)
            self.music_files_entry.insert(0, files)

    def add_advertisement(self):
        self.number = self.number + 1
        ad_frame = tk.Frame(self.ad_frame)
        ad_frame.pack(padx=5, pady=5)
        self.ad_frame_list.append(ad_frame)

        tk.Label(ad_frame, text=str(self.number) + ". Файл рекламы:").grid(row=0, column=0, padx=5, pady=5)
        ad_file_entry = tk.Entry(ad_frame, width=30)
        ad_file_entry.grid(row=0, column=1, padx=5, pady=5)
        dur_label = tk.Label(ad_frame, text='')
        dur_label.grid(row=0, column=5, padx=5, pady=5)
        tk.Button(ad_frame, text="Обзор", command=lambda: self.browse_ad_file(ad_file_entry, dur_label)).grid(row=0,
                                                                                                              column=2,
                                                                                                              padx=5,
                                                                                                              pady=5)

        tk.Label(ad_frame, text="Повторы:").grid(row=0, column=3, padx=5, pady=5)
        ad_repeat_entry = ttk.Combobox(ad_frame, values=["20", "15", "10", "5"], state="readonly")
        ad_repeat_entry.set("20")
        ad_repeat_entry.grid(row=0, column=4, padx=5, pady=5)

        ad_file_entry.bind("<KeyRelease>", lambda event: self.update_load())
        ad_repeat_entry.bind("<<ComboboxSelected>>", lambda event: self.update_load())

        self.ad_files.append(ad_file_entry)
        self.ad_repeat.append(ad_repeat_entry)
        self.update_load()

    def add_advertisement_init(self, path, repeat):
        self.number = self.number + 1
        ad_frame = tk.Frame(self.ad_frame)
        ad_frame.pack(padx=5, pady=5)
        self.ad_frame_list.append(ad_frame)

        tk.Label(ad_frame, text=str(self.number) + ". Файл рекламы:").grid(row=0, column=0, padx=5, pady=5)
        ad_file_entry = tk.Entry(ad_frame, width=30)
        ad_file_entry.grid(row=0, column=1, padx=5, pady=5)
        ad_file_entry.insert(0, path)
        dur_label = tk.Label(ad_frame, text='')
        dur_label.grid(row=0, column=5, padx=5, pady=5)
        tk.Button(ad_frame, text="Обзор", command=lambda: self.browse_ad_file(ad_file_entry, dur_label)).grid(row=0,
                                                                                                              column=2,
                                                                                                              padx=5,
                                                                                                              pady=5)

        tk.Label(ad_frame, text="Повторы:").grid(row=0, column=3, padx=5, pady=5)
        ad_repeat_entry = ttk.Combobox(ad_frame, values=["20", "15", "10", "5"], state="readonly")
        ad_repeat_entry.set(repeat)
        ad_repeat_entry.grid(row=0, column=4, padx=5, pady=5)

        ad_file_entry.bind("<KeyRelease>", lambda event: self.update_load())
        ad_repeat_entry.bind("<<ComboboxSelected>>", lambda event: self.update_load())

        self.ad_files.append(ad_file_entry)
        self.ad_repeat.append(ad_repeat_entry)
        self.update_load()

    def delete_advertisement(self):
        if len(self.ad_frame_list) == 0:
            messagebox.showerror("Ошибка", "Нет рекламы, которую можно удалить.")
        else:
            self.number = self.number - 1
            for widget in self.ad_frame_list[-1].winfo_children():
                widget.destroy()
            self.ad_frame_list[-1].destroy()
            self.ad_frame_list.pop()
            self.ad_repeat.pop()
            self.ad_files.pop()
            self.update_load()

    def browse_ad_file(self, ad_file_entry, dur_label):
        file = filedialog.askopenfilename(filetypes=[("Audio files", "*.mp3 *.wav *.flac *.m4a *.ogg")])
        if file:
            ad_file_entry.delete(0, tk.END)
            ad_file_entry.insert(0, file)
        dur_label.config(text='Продолжительность: ' + str(int(AudioSegment.from_file(file).duration_seconds)) + ' сек.')
        self.update_load()

    def generate_media_plan(self):
        ad_durations = []
        ad_repeats = []
        ad_names = []

        for ad_file_entry, ad_repeat_entry in zip(self.ad_files, self.ad_repeat):
            ad_file = ad_file_entry.get()
            ad_name = os.path.basename(ad_file)
            ad_dur = AudioSegment.from_file(ad_file).duration_seconds
            if isinstance(ad_repeat_entry, int):
                ad_repeat = ad_repeat_entry
            else:
                ad_repeat = ad_repeat_entry.get()

            if not ad_file or not ad_dur or not ad_repeat:
                messagebox.showerror("Ошибка", "Пожалуйста, заполните все поля для каждой рекламы.")
                return

            try:
                ad_dur = int(ad_dur)
                ad_repeat = int(ad_repeat)
            except ValueError:
                messagebox.showerror("Ошибка",
                                     "Продолжительность и количество повторов для рекламы должны быть целыми числами.")
                return

            ad_durations.append(ad_dur)
            ad_repeats.append(ad_repeat)
            ad_names.append(ad_name)

        self.ad_dur = ad_durations
        self.ad_repeat = ad_repeats
        self.ad_name = ad_names

        self.start_time = self.start_time_entry.get()
        self.end_time = self.end_time_entry.get()

        if not self.start_time or not self.end_time:
            messagebox.showerror("Ошибка", "Пожалуйста, введите начальное и конечное время.")
            return

        current_datetime = datetime.now().strftime("%Y-%m-%d %H-%M")

        # Инициализация переменных
        song_name = []
        song_dur = []

        # Обработка музыкальных файлов

        for file_path in os.listdir(self.music_files_entry.get()):
            filename = os.path.basename(file_path)
            try:
                song = AudioSegment.from_file(self.music_files_entry.get() + '/' + file_path).duration_seconds
                song_dur.append(song)
                song_name.append(filename)
            except Exception as e:
                print(f"Ошибка при обработке файла {filename}: {e}")
                # song_dur.append(None)

        # Преобразование времени
        object_time1 = hour_to_seconds(self.start_time)
        object_time2 = hour_to_seconds(self.end_time)

        indeces, self.ad_repeat = sort_list(self.ad_repeat)
        self.ad_name = rearrange(indeces, self.ad_name)
        self.ad_dur = rearrange(indeces, self.ad_dur)

        # Создание рабочей книги
        wb = openpyxl.Workbook()

        index_20_start = len(self.ad_name)
        for i in range(len(self.ad_name)):
            if self.ad_repeat[i] == 20:
                index_20_start = i
                break
        index_20_end = len(self.ad_name)
        index_20 = index_20_start

        if object_time2 < object_time1:
            working_time = 86400 + object_time2 - object_time1
        else:
            working_time = object_time2 - object_time1

        ad_sum = 0
        for k in range(len(self.ad_name)):
            ad_sum += self.ad_dur[k] * self.ad_repeat[k]
            percent = ad_sum / working_time
            adlimit = find_ab(percent)
        block_period = 0
        if len(self.ad_name) % adlimit == 0:
            block_period = len(self.ad_name) / adlimit
        else:
            block_period = int(len(self.ad_name) / adlimit) + 1

        blocks = int(block_period * 20)

        ad_uses = []
        ad_broadcast = []
        for i in range(len(self.ad_repeat)):
            ad_uses.append(0)
            ad_broadcast.append(0)
        period = int((working_time - 20) / blocks)
        cur_time = object_time1
        ad_start = object_time1 + period - 120
        cur_block = False
        ad_block = 1
        ads_in_block = 0
        timer = 0
        timestamps = []
        block_start = ''
        end_time = 0
        if object_time2 < object_time1:
            end_time = 86400 + object_time2
        else:
            end_time = object_time2

        self.load.config(text="Загружаемость: " + "{:.2f}".format(percent * 100) + "%")

        wb.create_sheet(str(int(working_time / 3600)))
        ws = wb[str(int(working_time / 3600))]
        ws["A1"] = "Имя"
        ws['C1'] = "Время"
        ws["D1"] = str(int(working_time / 3600)) + ' часа'
        ws["G1"] = 'Реклама'
        ws["H1"] = len(self.ad_name)
        ws["G2"] = "Повторы"
        ws["H2"] = "20:15:10:5"
        ws["G3"] = 'Продолжительность'
        ws["H3"] = sum(self.ad_dur)
        ws["G4"] = "Загруженность"
        ws["H4"] = "{:.2f}".format(percent * 100) + '%'
        ws["G5"] = "Максимальный рекламный блок"
        ws["H5"] = adlimit

        row_excel = 2
        colors = []
        green = PatternFill(patternType='solid', fgColor='78D542')
        colors.append(green)
        yellow = PatternFill(patternType='solid', fgColor='FFC638')
        colors.append(yellow)
        cur_color = 1

        while cur_time < end_time:
            block_start = cur_time
            while cur_block == False:
                cur_color = 1
                if cur_time > ad_start:
                    if int(ad_start) - block_start < 0:
                        col = "C"
                        while ws[col + str(row_excel - 1)].value is not None:
                            col = chr(ord(col) + 1)
                        ws[col + str(row_excel - 1)] = "00:00:00"
                        ws[chr(ord(col) + 1) + str(row_excel - 1)] = "Музыка"
                    else:
                        cur_time = ad_start
                        col = "C"
                        while ws[col + str(row_excel - 1)].value is not None:
                            col = chr(ord(col) + 1)
                        ws[col + str(row_excel - 1)] = sec_to_hour(int(cur_time) - block_start)
                        ws[chr(ord(col) + 1) + str(row_excel - 1)] = "Музыка"
                    ad_start += period
                    cur_block = True
                    continue
                rand_int = random.randint(0, len(song_name) - 1)
                ws["A" + str(row_excel)] = song_name[rand_int]
                ws["B" + str(row_excel)] = sec_to_hour(cur_time)

                ws['A' + str(row_excel)].fill = colors[cur_color]
                ws['B' + str(row_excel)].fill = colors[cur_color]
                row_excel += 1
                cur_time += int(float(song_dur[rand_int])) + 1

            while cur_block == True:
                cur_color = 0
                if (ad_block - 1) == blocks:
                    cur_block = False
                    continue
                timer = 0
                block_start = cur_time
                reset = 1
                if block_period == 1:
                    reset = 0
                if (ad_block) % block_period == reset:
                    index_20 = index_20_start
                for i in range(len(self.ad_repeat)):
                    if self.ad_repeat[i] == 5 and ad_uses[i] != self.ad_repeat[i]:
                        if (ad_block % int(blocks / 5) == 1) or (ad_broadcast[i] == 1):
                            if ads_in_block == adlimit:
                                ad_broadcast[i] = 1
                                continue
                            ad_broadcast[i] = 0
                            ws["A" + str(row_excel)] = self.ad_name[i]
                            ws["B" + str(row_excel)] = sec_to_hour(cur_time)
                            ws["C" + str(row_excel)] = ad_uses[i] + 1
                            ws['A' + str(row_excel)].fill = colors[cur_color]
                            ws['B' + str(row_excel)].fill = colors[cur_color]
                            row_excel += 1
                            ads_in_block += 1
                            ad_uses[i] += 1
                            timer += int(float(self.ad_dur[i])) + 1
                            cur_time += int(float(self.ad_dur[i])) + 1
                            continue

                    if self.ad_repeat[i] == 10 and ad_uses[i] != self.ad_repeat[i]:
                        if (ad_block % int(blocks / 10) == 1) or (ad_broadcast[i] == 1):
                            if ads_in_block == adlimit:
                                ad_broadcast[i] = 1
                                continue
                            ad_broadcast[i] = 0
                            ws["A" + str(row_excel)] = self.ad_name[i]
                            ws["B" + str(row_excel)] = sec_to_hour(cur_time)
                            ws["C" + str(row_excel)] = ad_uses[i] + 1
                            ws['A' + str(row_excel)].fill = colors[cur_color]
                            ws['B' + str(row_excel)].fill = colors[cur_color]
                            row_excel += 1
                            ads_in_block += 1
                            ad_uses[i] += 1
                            timer += int(float(self.ad_dur[i])) + 1
                            cur_time += int(float(self.ad_dur[i])) + 1
                            continue

                    if self.ad_repeat[i] == 15 and ad_uses[i] != self.ad_repeat[i]:
                        mod = 1
                        if int(blocks / 15) == 1:
                            mod = 0
                        if (ad_block % int(blocks / 15) == mod) or (ad_broadcast[i] == 1):
                            if ads_in_block == adlimit:
                                ad_broadcast[i] = 1
                                continue
                            ad_broadcast[i] = 0
                            ws["A" + str(row_excel)] = self.ad_name[i]
                            ws["B" + str(row_excel)] = sec_to_hour(cur_time)
                            ws["C" + str(row_excel)] = ad_uses[i] + 1
                            ws['A' + str(row_excel)].fill = colors[cur_color]
                            ws['B' + str(row_excel)].fill = colors[cur_color]
                            row_excel += 1
                            ads_in_block += 1
                            ad_uses[i] += 1
                            timer += int(float(self.ad_dur[i])) + 1
                            cur_time += int(float(self.ad_dur[i])) + 1
                            continue

                    if self.ad_repeat[i] == 20 and ad_uses[i] != self.ad_repeat[i]:
                        if (ads_in_block == adlimit or i != index_20):
                            continue
                        ws["A" + str(row_excel)] = self.ad_name[i]
                        ws["B" + str(row_excel)] = sec_to_hour(cur_time)
                        ws["C" + str(row_excel)] = ad_uses[i] + 1
                        ws['A' + str(row_excel)].fill = colors[cur_color]
                        ws['B' + str(row_excel)].fill = colors[cur_color]
                        row_excel += 1
                        ads_in_block += 1
                        ad_uses[i] += 1
                        index_20 += 1
                        timer += int(float(self.ad_dur[i])) + 1
                        cur_time += int(float(self.ad_dur[i])) + 1
                        continue

                col = "C"
                while ws[col + str(row_excel - 1)].value is not None:
                    col = chr(ord(col) + 1)
                ws[col + str(row_excel - 1)] = sec_to_hour(int(cur_time) - block_start)
                ws[chr(ord(col) + 1) + str(row_excel - 1)] = "Реклама"
                ads_in_block = 0
                ad_block += 1
                cur_block = False

        ws.append([''])
        ws.append([''])
        ws.append(['Повторов за день'])
        for row in zip(ad_uses, self.ad_name):
            ws.append(row)

        ws.column_dimensions['A'].width = 66

        filename = f"Медиаплан {len(self.ad_name)} реклам, средняя прод. {sum(self.ad_dur) / len(self.ad_name)} сек, раб. время {int(working_time / 3600)} ч., {current_datetime}.xlsx"

        del wb['Sheet']
        wb.save(filename)
        messagebox.showinfo("Успех", f"Медиаплан успешно создан и сохранен под именем {filename}!")

    def update_load(self):
        if not self.start_time_entry.get() or not self.end_time_entry.get():
            return

        ad_durations = []
        ad_repeats = []

        for ad_file_entry, ad_repeat_entry in zip(self.ad_files, self.ad_repeat):
            ad_file = ad_file_entry.get()
            ad_dur = AudioSegment.from_file(ad_file).duration_seconds if os.path.exists(ad_file) else 0
            if isinstance(ad_repeat_entry, ttk.Combobox):
                ad_repeat = ad_repeat_entry.get()
            else:
                ad_repeat = ad_repeat_entry

            try:
                ad_dur = int(ad_dur)
                ad_repeat = int(ad_repeat)
            except ValueError:
                return

            ad_durations.append(ad_dur)
            ad_repeats.append(ad_repeat)

        start_time = hour_to_seconds(self.start_time_entry.get())
        end_time = hour_to_seconds(self.end_time_entry.get())

        if end_time < start_time:
            working_time = 86400 + end_time - start_time
        else:
            working_time = end_time - start_time

        ad_sum = sum(dur * repeat for dur, repeat in zip(ad_durations, ad_repeats))
        percent = ad_sum / working_time
        self.load.config(text="Загружаемость: " + "{:.2f}".format(percent * 100) + "%")

    def save_json(self):
        data = {
            "music": self.music_files_entry.get(),
            "start": self.start_time_entry.get(),
            "end": self.end_time_entry.get()
        }

        if len(self.ad_frame_list) != 0:
            ads = []
            for i in range(len(self.ad_frame_list)):
                ad_info = []
                ad_info.append(self.ad_files[i].get())
                if (isinstance(self.ad_repeat[i], int)):
                    ad_info.append(str(self.ad_repeat[i]))
                else:
                    ad_info.append(self.ad_repeat[i].get())
                ads.append(ad_info)
            data.update({"ad": ads})
        exe_path = os.path.abspath(sys.argv[0])
        exe_dir = os.path.dirname(exe_path)

        json_object = json.dumps(data, indent=4)
        with open(exe_dir + "/data.json", "w", encoding='ascii') as outfile:
            outfile.write(json_object)

        self.root.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    app = MediaPlanApp(root)
    root.protocol("WM_DELETE_WINDOW", app.save_json)
    root.mainloop()
