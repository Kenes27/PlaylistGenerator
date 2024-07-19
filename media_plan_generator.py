import os
import random
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
from pydub import AudioSegment
from tkinter import messagebox
from utils import find_ab, hour_to_seconds, sec_to_hour, sort_list, rearrange

class MediaPlanGenerator:

    def __init__(self, parent):
        self.parent = parent
        self.ad_durations = []
        self.ad_repeats = []
        self.ad_names = []
        self.start_time = parent.start_time_entry.get()
        self.end_time = parent.end_time_entry.get()

    def collect_ad_data(self):
        for ad_file_entry, ad_repeat_entry in zip(self.parent.ad_files, self.parent.ad_repeat_combo):
            ad_file = ad_file_entry.get()
            ad_name = os.path.basename(ad_file)
            ad_dur = AudioSegment.from_file(ad_file).duration_seconds
            ad_repeat = ad_repeat_entry.get()

            if not ad_file or not ad_dur or not ad_repeat:
                messagebox.showerror("Ошибка", "Пожалуйста, заполните все поля для каждой рекламы.")
                return False

            try:
                ad_dur = int(ad_dur)
                ad_repeat = int(ad_repeat)
            except ValueError:
                messagebox.showerror("Ошибка", "Продолжительность и количество повторов для рекламы должны быть целыми числами.")
                return False

            self.ad_durations.append(ad_dur)
            self.ad_repeats.append(ad_repeat)
            self.ad_names.append(ad_name)

        return True

    def collect_music_data(self):
        self.song_name = []
        self.song_dur = []
        music_folder = self.parent.music_files_entry.get()

        if not os.path.isdir(music_folder):
            messagebox.showerror("Ошибка", "Пожалуйста, укажите папку с музыкальными файлами.")
            return False

        for file_path in os.listdir(music_folder):
            filename = os.path.basename(file_path)
            try:
                song = AudioSegment.from_file(os.path.join(music_folder, file_path)).duration_seconds
                self.song_dur.append(song)
                self.song_name.append(filename)
            except Exception as e:
                print(f"Ошибка при обработке файла {filename}: {e}")
                messagebox.showerror("Ошибка", f"Ошибка при обработке файла {filename}: {e}")
                return False

        return True

    def generate(self):
        if not self.collect_ad_data():
            return

        if not self.collect_music_data():
            return

        current_datetime = datetime.now().strftime("%Y-%m-%d %H-%M")
        object_time1 = hour_to_seconds(self.start_time)
        object_time2 = hour_to_seconds(self.end_time)

        indeces, self.ad_repeats = sort_list(self.ad_repeats)
        self.ad_names = rearrange(indeces, self.ad_names)
        self.ad_durations = rearrange(indeces, self.ad_durations)

        if object_time2 < object_time1:
            working_time = 86400 + object_time2 - object_time1
        else:
            working_time = object_time2 - object_time1

        ad_sum = sum(dur * repeat for dur, repeat in zip(self.ad_durations, self.ad_repeats))
        percent = ad_sum / working_time
        adlimit = find_ab(percent)
        block_period = (len(self.ad_names) + adlimit - 1) // adlimit
        blocks = int(block_period * 20)

        ad_uses = []
        ad_broadcast = []
        for i in range(len(self.ad_repeats)):
            ad_uses.append(0)
            ad_broadcast.append(0)
        period = int((working_time - 20) / blocks)
        cur_time = object_time1
        ad_start = object_time1 + period - 120
        cur_block = False
        ad_block = 1
        ads_in_block = 0
        timer = 0
        timestamps = []
        block_start = ''
        end_time = 0
        if object_time2 < object_time1:
            end_time = 86400 + object_time2
        else:
            end_time = object_time2

        self.parent.load.config(text="Загружаемость: " + "{:.2f}".format(percent * 100) + "%")

        wb = openpyxl.Workbook()
        index_20_start = len(self.ad_names)
        for i in range(len(self.ad_names)):
            if self.ad_repeats[i] == 20:
                index_20_start = i
                break
        index_20_end = len(self.ad_names)
        index_20 = index_20_start

        wb.create_sheet(str(int(working_time / 3600)))
        ws = wb[str(int(working_time / 3600))]
        ws["A1"] = "Имя"
        ws['C1'] = "Время"
        ws["D1"] = str(int(working_time / 3600)) + ' часа'
        ws["G1"] = 'Реклама'
        ws["H1"] = len(self.ad_names)
        ws["G2"] = "Повторы"
        ws["H2"] = "20:15:10:5"
        ws["G3"] = 'Продолжительность'
        ws["H3"] = sum(self.ad_durations)
        ws["G4"] = "Загруженность"
        ws["H4"] = "{:.2f}".format(percent * 100) + '%'
        ws["G5"] = "Максимальный рекламный блок"
        ws["H5"] = adlimit

        row_excel = 2
        colors = []
        green = PatternFill(patternType='solid', fgColor='78D542')
        colors.append(green)
        yellow = PatternFill(patternType='solid', fgColor='FFC638')
        colors.append(yellow)
        cur_color = 1

        while cur_time < end_time:
            block_start = cur_time
            while cur_block == False:
                cur_color = 1
                if cur_time > ad_start:
                    if int(ad_start) - block_start < 0:
                        col = "C"
                        while ws[col + str(row_excel - 1)].value is not None:
                            col = chr(ord(col) + 1)
                        ws[col + str(row_excel - 1)] = "00:00:00"
                        ws[chr(ord(col) + 1) + str(row_excel - 1)] = "Музыка"
                    else:
                        cur_time = ad_start
                        col = "C"
                        while ws[col + str(row_excel - 1)].value is not None:
                            col = chr(ord(col) + 1)
                        ws[col + str(row_excel - 1)] = sec_to_hour(int(cur_time) - block_start)
                        ws[chr(ord(col) + 1) + str(row_excel - 1)] = "Музыка"
                    ad_start += period
                    cur_block = True
                    continue
                rand_int = random.randint(0, len(self.song_name) - 1)
                ws["A" + str(row_excel)] = self.song_name[rand_int]
                ws["B" + str(row_excel)] = sec_to_hour(cur_time)

                ws['A' + str(row_excel)].fill = colors[cur_color]
                ws['B' + str(row_excel)].fill = colors[cur_color]
                row_excel += 1
                cur_time += int(float(self.song_dur[rand_int])) + 1

            while cur_block == True:
                cur_color = 0
                if (ad_block - 1) == blocks:
                    cur_block = False
                    continue
                timer = 0
                block_start = cur_time
                reset = 1
                if block_period == 1:
                    reset = 0
                if (ad_block) % block_period == reset:
                    index_20 = index_20_start
                for i in range(len(self.ad_repeats)):
                    if self.ad_repeats[i] == 5 and ad_uses[i] != self.ad_repeats[i]:
                        if (ad_block % int(blocks / 5) == 1) or (ad_broadcast[i] == 1):
                            if ads_in_block == adlimit:
                                ad_broadcast[i] = 1
                                continue
                            ad_broadcast[i] = 0
                            ws["A" + str(row_excel)] = self.ad_names[i]
                            ws["B" + str(row_excel)] = sec_to_hour(cur_time)
                            ws["C" + str(row_excel)] = ad_uses[i] + 1
                            ws['A' + str(row_excel)].fill = colors[cur_color]
                            ws['B' + str(row_excel)].fill = colors[cur_color]
                            row_excel += 1
                            ads_in_block += 1
                            ad_uses[i] += 1
                            timer += int(float(self.ad_durations[i])) + 1
                            cur_time += int(float(self.ad_durations[i])) + 1
                            continue

                    if self.ad_repeats[i] == 10 and ad_uses[i] != self.ad_repeats[i]:
                        if (ad_block % int(blocks / 10) == 1) or (ad_broadcast[i] == 1):
                            if ads_in_block == adlimit:
                                ad_broadcast[i] = 1
                                continue
                            ad_broadcast[i] = 0
                            ws["A" + str(row_excel)] = self.ad_names[i]
                            ws["B" + str(row_excel)] = sec_to_hour(cur_time)
                            ws["C" + str(row_excel)] = ad_uses[i] + 1
                            ws['A' + str(row_excel)].fill = colors[cur_color]
                            ws['B' + str(row_excel)].fill = colors[cur_color]
                            row_excel += 1
                            ads_in_block += 1
                            ad_uses[i] += 1
                            timer += int(float(self.ad_durations[i])) + 1
                            cur_time += int(float(self.ad_durations[i])) + 1
                            continue

                    if self.ad_repeats[i] == 15 and ad_uses[i] != self.ad_repeats[i]:
                        mod = 1
                        if int(blocks / 15) == 1:
                            mod = 0
                        if (ad_block % int(blocks / 15) == mod) or (ad_broadcast[i] == 1):
                            if ads_in_block == adlimit:
                                ad_broadcast[i] = 1
                                continue
                            ad_broadcast[i] = 0
                            ws["A" + str(row_excel)] = self.ad_names[i]
                            ws["B" + str(row_excel)] = sec_to_hour(cur_time)
                            ws["C" + str(row_excel)] = ad_uses[i] + 1
                            ws['A' + str(row_excel)].fill = colors[cur_color]
                            ws['B' + str(row_excel)].fill = colors[cur_color]
                            row_excel += 1
                            ads_in_block += 1
                            ad_uses[i] += 1
                            timer += int(float(self.ad_durations[i])) + 1
                            cur_time += int(float(self.ad_durations[i])) + 1
                            continue

                    if self.ad_repeats[i] == 20 and ad_uses[i] != self.ad_repeats[i]:
                        if (ads_in_block == adlimit or i != index_20):
                            continue
                        ws["A" + str(row_excel)] = self.ad_names[i]
                        ws["B" + str(row_excel)] = sec_to_hour(cur_time)
                        ws["C" + str(row_excel)] = ad_uses[i] + 1
                        ws['A' + str(row_excel)].fill = colors[cur_color]
                        ws['B' + str(row_excel)].fill = colors[cur_color]
                        row_excel += 1
                        ads_in_block += 1
                        ad_uses[i] += 1
                        index_20 += 1
                        timer += int(float(self.ad_durations[i])) + 1
                        cur_time += int(float(self.ad_durations[i])) + 1
                        continue

                col = "C"
                while ws[col + str(row_excel - 1)].value is not None:
                    col = chr(ord(col) + 1)
                ws[col + str(row_excel - 1)] = sec_to_hour(int(cur_time) - block_start)
                ws[chr(ord(col) + 1) + str(row_excel - 1)] = "Реклама"
                ads_in_block = 0
                ad_block += 1
                cur_block = False

        ws.append([''])
        ws.append([''])
        ws.append(['Повторов за день'])
        for row in zip(ad_uses, self.ad_names):
            ws.append(row)

        ws.column_dimensions['A'].width = 66

        average_duration = round(sum(self.ad_durations) / len(self.ad_names), 1)
        filename = f"Медиаплан {len(self.ad_names)} реклам, средняя прод. {average_duration} сек, раб. время {int(working_time / 3600)} ч., {current_datetime}.xlsx"

        del wb['Sheet']
        wb.save(filename)
        messagebox.showinfo("Успех", f"Медиаплан успешно создан и сохранен под именем {filename}!")
